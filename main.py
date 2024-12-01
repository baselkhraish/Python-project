import openpyxl
import pickle
from customer import Customer

workbook = openpyxl.load_workbook('file.xlsx')
sheet = workbook.active
# 1
for row in sheet.iter_rows(min_row=1, values_only=True):
    print(row)

# 4
customers = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    customer = Customer(*row)
    customers.append(customer)

for customer in customers:
    print(customer)

# 5
for customer in customers:
    print(f"ID: {customer.ID}, Year of Birth: {customer.Year_Birth}, Income: {customer.Income}")

# 6
total_income = 0
total_customers = len(customers)

for customer in customers:
    income = customer.Income
    if(income == None):
        total_customers -= 1
    if (income != None):
        total_income += income

average_income = total_income / total_customers if total_customers > 0 else 0

print(f"The average income of all customers is: ${average_income:,.2f}")


# 7
def top_customers(customers, average_income):
    top_customers_list = []

    for customer in customers:
        income = 0
        if (customer.Income != None):
            income += float(customer.Income)

        if income > average_income:
            top_customers_list.append(customer)

    return top_customers_list

top_customers_list = top_customers(customers, average_income)

for customer in top_customers_list:
    print(f"ID: {customer.ID}, Income: {customer.Income}")


# 8
def countMarried(customers):
    married_count = 0
    for customer in customers:
        if customer.Marital_Status == "Married":
            married_count += 1
    return married_count

married_customers_count = countMarried(customers)

print(f"The number of customers who are married: {married_customers_count}")

# 9
def unique_education_degrees(customers):
    education_degrees = set()

    for customer in customers:
        education_degrees.add(customer.Education)

    return education_degrees

unique_degrees = unique_education_degrees(customers)

print("Unique educational degrees:")
for degree in unique_degrees:
    print(degree)

# 10
successful_customers_count = sum(1 for customer in customers if customer.Successful_Campaigns == 1)

print(f"The number of successful customers is: {successful_customers_count}")


# 11
def classifyEducationalDegrees(customers):
    education_count = {}

    for customer in customers:
        degree = customer.Education

        if degree in education_count:
            education_count[degree] += 1
        else:
            education_count[degree] = 1

    return education_count

education_summary = classifyEducationalDegrees(customers)

print("Classified educational degrees and their counts:")
for degree, count in education_summary.items():
    print(f"{degree}: {count}")

# 12
average_income_cell = sheet.cell(row=2, column=5)
average_income_cell.value = average_income

workbook.save('file.xlsx')

print(f"The average income has been written to the Excel file: ${average_income:,.2f}")

# 13
with open('customers.pkl', 'wb') as f:
    pickle.dump(customers, f)

print("The list of customers has been saved to 'customers.pkl'.")

with open('customers.pkl', 'rb') as f:
    loaded_customers = pickle.load(f)

for customer in loaded_customers[:5]:
    print(customer)
